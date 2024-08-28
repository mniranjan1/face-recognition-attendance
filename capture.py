from __future__ import print_function
import cv2
import mediapipe as mp
import numpy as np
import time
import imutils
import os
import fnmatch
from PIL import Image
from datetime import datetime
import smtplib
import openpyxl
from pygame import mixer
import utils
from facenet_pytorch import MTCNN, InceptionResnetV1
import torch
from motion_detection import motion_detection


# Check if CUDA is available
if torch.cuda.is_available():
    device = torch.device("cuda")
    print("CUDA is available. Using GPU:", torch.cuda.get_device_name(0))
else:
    device = torch.device("cpu")
    print("CUDA is not available. Using CPU.")


def find(pattern, path):
    result = []
    for root, _, files in os.walk(path):
        for name in files:
            if fnmatch.fnmatch(name, pattern):
                result.append(os.path.join(root, name))
    return result


# Starting the mixer
mixer.init()
mixer.music.load("door_bell.mp3")
mixer.music.set_volume(0.7)

modelFile = "models/res10_300x300_ssd_iter_140000.caffemodel"
configFile = "models/deploy.prototxt.txt"
net = cv2.dnn.readNetFromCaffe(configFile, modelFile)


def takeAttendence(name):
    now = datetime.now()
    month = now.month
    file = str(month) + ".xlsx"
    path = "E:\\Projects\\test\\Face_recogntion_final"
    if find(file, path):
        book = openpyxl.load_workbook(file)
        sheet = book.active
    else:
        book = openpyxl.Workbook()
        sheet = book.active
        headers = ["Name", "Department", "Checkin", "Checkout"]
        sheet.append(headers)
    max = sheet.max_row
    data1 = []
    d = datetime.strptime("15:30", "%H:%M")
    timestring = now.strftime("%H:%M")
    x1 = sheet["A"]
    x1 = x1[1:]
    for i in x1:
        if i.value != name:
            if now.time() < d.time():
                data1.extend([name, "AI", timestring])
                for col, entry in enumerate(data1, start=1):
                    sheet.cell(row=max + 1, column=col, value=entry)
            else:
                data1.extend([timestring])
                for col, entry in enumerate(data1, start=4):
                    sheet.cell(row=max + 1, column=col, value=entry)
        else:
            break

    book.save(str(month) + ".xlsx")

    EMAIL = "source_email_address"
    PASS = "pass_key"
    REMAIL = "dest_email_address"
    s = smtplib.SMTP("smtp.gmail.com:587")
    s.ehlo()
    s.starttls()
    s.login(EMAIL, PASS)
    message = name + "   Your attendance has been done."
    s.sendmail(EMAIL, REMAIL, message)
    s.quit()


def start_program():
    face_recog = cv2.imread("face_recog.png")
    cv2.putText(
        face_recog,
        "Welcome to Face Recognition",
        (150, 100),
        cv2.FONT_HERSHEY_SIMPLEX,
        1,
        (0, 0, 255),
        2,
    )
    cv2.putText(
        face_recog,
        "Attendance System",
        (150, 140),
        cv2.FONT_HERSHEY_SIMPLEX,
        1,
        (0, 0, 255),
        2,
    )
    cv2.putText(
        face_recog,
        "1. Press 'S' to Start",
        (150, 500),
        cv2.FONT_HERSHEY_SIMPLEX,
        1,
        (0, 0, 255),
        2,
    )
    cv2.putText(
        face_recog,
        "2. Press 'Q' to Exit",
        (150, 550),
        cv2.FONT_HERSHEY_SIMPLEX,
        1,
        (0, 0, 255),
        2,
    )

    cv2.imshow("img", face_recog)

    Camera = False  # Initialize Camera to False

    while True:
        k = cv2.waitKey(33) & 0xFF  # Wait for a key press for 33 milliseconds

        if k == ord("s"):  # Start when 'S' is pressed
            print("S key pressed, starting motion detection...")
            camera_on = (
                motion_detection()
            )  # Assuming this is a function you have defined
            if camera_on:
                Camera = True
            return Camera
        elif k == ord("q"):  # Exit when 'Q' is pressed
            return False


def check_face(Camera):
    T = 0
    F = 0
    Attendance = False
    detect = True
    sentmail = False
    pred_name = ""
    Capture_image = False
    unknown = False
    mevm, cls_names = utils.data_load()

    # Define Inception Resnet V1 module (GoogLe Net)

    if Camera == True:
        resnet = (
            InceptionResnetV1(pretrained="vggface2", num_classes=10).eval().to(device)
        )
        mtcnn = MTCNN(
            image_size=160,
            margin=0,
            min_face_size=20,
            thresholds=[0.6, 0.7, 0.7],
            factor=0.709,
            device=device,
            keep_all=True,
        )
        video = 0
        cap = cv2.VideoCapture(video)
        # Facetection class to be loaded
        while cap.isOpened():
            ret, frame0 = cap.read()

            # timer =time.time()
            if Attendance == False and detect == True:
                frame0 = imutils.resize(frame0, width=700)
                frame = cv2.cvtColor(frame0, cv2.COLOR_BGR2RGB)
                frames = Image.fromarray(frame)
                img_cropped = mtcnn(frames)
                height, width = frame0.shape[:2]
                blob = cv2.dnn.blobFromImage(
                    cv2.resize(frame0, (300, 300)),
                    1.0,
                    (300, 300),
                    (104.0, 117.0, 123.0),
                )
                net.setInput(blob)
                faces3 = net.forward()
                # OPENCV DNN
                for i in range(faces3.shape[2]):
                    confidence = faces3[0, 0, i, 2]
                    if confidence > 0.5:
                        box = faces3[0, 0, i, 3:7] * np.array(
                            [width, height, width, height]
                        )
                        (x, y, x1, y1) = box.astype("int")
                        cv2.rectangle(frame0, (x, y), (x1, y1), (0, 0, 255), 2)
                try:
                    img_embedding = resnet(img_cropped.to(device))
                    text, frame_final = utils.calc_dist(
                        img_embedding.detach().cpu().numpy(),
                        mevm,
                        cls_names,
                        frame0,
                        x,
                        y,
                    )
                    if text == "unknown":
                        F += 1
                        # print("Person not matched")
                        # cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 255), 4)
                        if F > 3:
                            unknown = True
                            Attendance = False
                            detect = False
                    else:
                        T += 1
                        if T > 100:
                            pred_name = text
                            # print(pred_name)
                            Attendance = True
                            detect = False
                except Exception as e:
                    print(e)
                    frame_final = frame0

            elif Attendance == True and detect == False:
                # takeAttendence(pred_name)
                detect = True
                sentmail = True

            elif sentmail == True:
                cv2.putText(
                    frame_final,
                    pred_name
                    + "!!!"
                    + " Your attendance has been done!!!!\n press 'Q' to exit",
                    (10, 20),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.7,
                    (0, 0, 255),
                    2,
                )

            elif unknown == True:
                cv2.putText(
                    frame_final,
                    "Your image doesn't exist in our database",
                    (10, 20),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.7,
                    (255, 0, 0),
                    2,
                )
                cv2.putText(
                    frame_final,
                    "press 'Y' to start capturing images or press 'Q' to exit",
                    (10, 50),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.7,
                    (255, 0, 0),
                    2,
                )

            cv2.imshow("Output", frame_final)
            # Save Woorksheet as present month
            key = cv2.waitKey(1)
            if key == 113:  # "q"
                break
            if key == ord("y"):
                Capture_image = True
                break
    cap.release()
    cv2.destroyAllWindows()
    return Capture_image


def init_create_folder():
    # create the folder and database if not exist
    if not os.path.exists("images"):
        os.mkdir("images")


def create_folder(folder_name):
    if not os.path.exists(folder_name):
        os.mkdir(folder_name)


def store_images(img_name):
    mp_face_mesh = mp.solutions.face_mesh
    face_mesh = mp_face_mesh.FaceMesh(
        min_detection_confidence=0.5, min_tracking_confidence=0.5
    )

    mp_drawing = mp.solutions.drawing_utils

    drawing_spec = mp_drawing.DrawingSpec(thickness=1, circle_radius=1)

    cap = cv2.VideoCapture(0)
    flag_start_capturing = False
    frames = 0
    total_pics = 200
    pic_no = 0
    F, L, R, U, D = 0, 0, 0, 0, 0
    create_folder("images/" + str(img_name))
    Forward = True
    Left = True
    Down = True
    Up = True
    Right = True
    show = True
    Capture = True
    Camera = True

    while cap.isOpened():
        success, image = cap.read()

        image = imutils.resize(image, width=800)

        image = cv2.flip(image, 1)

        start = time.time()
        if Camera == True:
            cv2.putText(
                image,
                "Press 'C' to start Capturing",
                (10, 20),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.6,
                (255, 0, 0),
                2,
            )

            cv2.putText(
                image,
                "Press 'Q' to quit",
                (10, 40),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.6,
                (0, 0, 255),
                2,
            )

        if frames > 1 and Capture == True:
            Camera = False
            text2 = "Starting Capturing....."
            text3 = "Please be patient"
            cv2.putText(
                image, text2, (200, 150), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 0), 2
            )
            cv2.putText(
                image, text3, (200, 300), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 0), 2
            )

        # Flip the image horizontally for a later selfie-view display
        # Also convert the color space from BGR to RGB
        if frames > 150:
            Capture = False

            clone = image.copy()

            image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
            # To improve performance
            image.flags.writeable = False

            # Get the result
            results = face_mesh.process(image)

            # To improve performance
            image.flags.writeable = True

            # Convert the color space from RGB to BGR
            image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)

            img_h, img_w, img_c = image.shape
            face_3d = []
            face_2d = []

            if results.multi_face_landmarks:
                for face_landmarks in results.multi_face_landmarks:
                    for idx, lm in enumerate(face_landmarks.landmark):
                        if (
                            idx == 33
                            or idx == 263
                            or idx == 1
                            or idx == 61
                            or idx == 291
                            or idx == 199
                        ):
                            if idx == 1:
                                nose_2d = (lm.x * img_w, lm.y * img_h)
                                nose_3d = (lm.x * img_w, lm.y * img_h, lm.z * 3000)

                            x, y = int(lm.x * img_w), int(lm.y * img_h)

                            # Get the 2D Coordinates
                            face_2d.append([x, y])

                            # Get the 3D Coordinates
                            face_3d.append([x, y, lm.z])

                    # Convert it to the NumPy array
                    face_2d = np.array(face_2d, dtype=np.float64)

                    # Convert it to the NumPy array
                    face_3d = np.array(face_3d, dtype=np.float64)

                    # The camera matrix
                    focal_length = 1 * img_w

                    cam_matrix = np.array(
                        [
                            [focal_length, 0, img_h / 2],
                            [0, focal_length, img_w / 2],
                            [0, 0, 1],
                        ]
                    )

                    # The distortion parameters
                    dist_matrix = np.zeros((4, 1), dtype=np.float64)

                    # Solve PnP
                    success, rot_vec, trans_vec = cv2.solvePnP(
                        face_3d, face_2d, cam_matrix, dist_matrix
                    )

                    # Get rotational matrix
                    rmat, jac = cv2.Rodrigues(rot_vec)

                    # Get angles
                    angles, mtxR, mtxQ, Qx, Qy, Qz = cv2.RQDecomp3x3(rmat)

                    # Get the y rotation degree
                    x = angles[0] * 360
                    y = angles[1] * 360
                    z = angles[2] * 360

                    # See where the user's head tilting
                    if y < -10:
                        text = "Looking Left"
                        L += 1
                    elif y > 10:
                        text = "Looking Right"
                        R += 1
                    elif x < -10:
                        text = "Looking Down"
                        D += 1
                    elif x > 10:
                        text = "Looking Up"
                        U += 1
                    else:
                        text = "Forward"
                        F += 1
                        if show == True:
                            text1 = "Please Look Straight Forward"
                            cv2.putText(
                                image,
                                text1,
                                (20, 150),
                                cv2.FONT_HERSHEY_SIMPLEX,
                                1,
                                (255, 0, 0),
                                2,
                            )

                    if F > 100 and Forward == True:
                        if pic_no != total_pics:
                            cv2.imwrite(
                                "images/" + str(img_name) + "/" + str(pic_no) + ".jpg",
                                clone,
                            )
                            pic_no += 1
                            st = int((pic_no) * 100 / total_pics)
                            # draw the segmented hand
                            cv2.putText(
                                image,
                                "Capturing..." + str(st) + " %",
                                (200, 50),
                                cv2.FONT_HERSHEY_SIMPLEX,
                                1,
                                (0, 0, 255),
                                2,
                            )
                            music_play = True
                        else:
                            if music_play == True:
                                mixer.music.play()
                            show = False
                            music_play = False
                            text1 = "Please Look Left"
                            cv2.arrowedLine(
                                image,
                                (400, 140),
                                (300, 140),
                                (255, 0, 0),
                                3,
                                cv2.LINE_AA,
                                0,
                                0.3,
                            )
                            cv2.putText(
                                image,
                                text1,
                                (20, 150),
                                cv2.FONT_HERSHEY_SIMPLEX,
                                1,
                                (255, 0, 0),
                                2,
                            )

                    if L > 50 and Left == True:
                        Forward = False
                        if pic_no != (total_pics * 2):
                            cv2.imwrite(
                                "images/" + str(img_name) + "/" + str(pic_no) + ".jpg",
                                clone,
                            )
                            pic_no += 1
                            st = int((pic_no) * 100 / (total_pics * 2))
                            # draw the segmented hand
                            cv2.putText(
                                image,
                                "Capturing..." + str(st) + " %",
                                (200, 50),
                                cv2.FONT_HERSHEY_SIMPLEX,
                                1,
                                (0, 0, 255),
                                2,
                            )
                            music_play = True
                        else:
                            if music_play == True:
                                mixer.music.play()
                            music_play = False
                            text1 = "Please Look Right"
                            cv2.arrowedLine(
                                image,
                                (350, 140),
                                (400, 140),
                                (255, 0, 0),
                                3,
                                cv2.LINE_AA,
                                0,
                                0.3,
                            )
                            cv2.putText(
                                image,
                                text1,
                                (20, 150),
                                cv2.FONT_HERSHEY_SIMPLEX,
                                1,
                                (255, 0, 0),
                                2,
                            )

                    if R > 50 and Right == True:
                        Left = False
                        if pic_no != (total_pics * 3):
                            cv2.imwrite(
                                "images/" + str(img_name) + "/" + str(pic_no) + ".jpg",
                                clone,
                            )
                            pic_no += 1
                            st = int((pic_no) * 100 / (total_pics * 3))
                            # draw the segmented hand
                            cv2.putText(
                                image,
                                "Capturing..." + str(st) + " %",
                                (200, 50),
                                cv2.FONT_HERSHEY_SIMPLEX,
                                1,
                                (0, 0, 255),
                                2,
                            )
                            music_play = True
                        else:
                            if music_play == True:
                                mixer.music.play()
                            music_play = False
                            text1 = "Please Look Up"
                            cv2.arrowedLine(
                                image,
                                (300, 140),
                                (300, 70),
                                (255, 0, 0),
                                3,
                                cv2.LINE_AA,
                                0,
                                0.3,
                            )
                            cv2.putText(
                                image,
                                text1,
                                (20, 150),
                                cv2.FONT_HERSHEY_SIMPLEX,
                                1,
                                (255, 0, 0),
                                2,
                            )

                    if U > 50 and Up == True:
                        Right = False
                        if pic_no != (total_pics * 4):
                            cv2.imwrite(
                                "images/" + str(img_name) + "/" + str(pic_no) + ".jpg",
                                clone,
                            )
                            pic_no += 1
                            st = int((pic_no) * 100 / (total_pics * 4))
                            # draw the segmented hand
                            cv2.putText(
                                image,
                                "Capturing..." + str(st) + " %",
                                (200, 50),
                                cv2.FONT_HERSHEY_SIMPLEX,
                                1,
                                (0, 0, 255),
                                2,
                            )
                            music_play = True
                        else:
                            if music_play == True:
                                mixer.music.play()
                            music_play = False
                            text1 = "Please Look Down"
                            cv2.arrowedLine(
                                image,
                                (300, 140),
                                (300, 210),
                                (255, 0, 0),
                                3,
                                cv2.LINE_AA,
                                0,
                                0.3,
                            )
                            cv2.putText(
                                image,
                                text1,
                                (20, 150),
                                cv2.FONT_HERSHEY_SIMPLEX,
                                1,
                                (255, 0, 0),
                                2,
                            )

                    if D > 50 and Down == True:
                        Up = False
                        if pic_no != (total_pics * 5):
                            cv2.imwrite(
                                "images/" + str(img_name) + "/" + str(pic_no) + ".jpg",
                                clone,
                            )
                            pic_no += 1
                            st = int((pic_no) * 100 / (total_pics * 5))
                            # draw the segmented hand
                            cv2.putText(
                                image,
                                "Capturing..." + str(st) + " %",
                                (200, 50),
                                cv2.FONT_HERSHEY_SIMPLEX,
                                1,
                                (0, 0, 255),
                                2,
                            )
                            music_play = True
                        else:
                            if music_play == True:
                                mixer.music.play()
                            music_play = False
                            Down = False
                            # cv2.putText(image, 'Youre image has been successfully Captured. You may now exit', (20, 150), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 0), 2)

                    # Display the nose direction
                    nose_3d_projection, jacobian = cv2.projectPoints(
                        nose_3d, rot_vec, trans_vec, cam_matrix, dist_matrix
                    )

                    p1 = (int(nose_2d[0]), int(nose_2d[1]))
                    p2 = (int(nose_2d[0] + y * 10), int(nose_2d[1] - x * 10))

                    # cv2.line(image, p1, p2, (255, 0, 0), 3)
                    # Add the text on the image
                    if Down == True:
                        cv2.putText(
                            image,
                            text,
                            (20, 50),
                            cv2.FONT_HERSHEY_SIMPLEX,
                            1,
                            (0, 255, 0),
                            2,
                        )
                    else:
                        cv2.putText(
                            image,
                            "You're image has been successfully Captured. You may now press 'Q' to exit!!! ",
                            (10, 20),
                            cv2.FONT_HERSHEY_SIMPLEX,
                            0.6,
                            (255, 0, 0),
                            2,
                        )

            end = time.time()
            totalTime = end - start

        cv2.imshow("Image Capture", image)

        keypress = cv2.waitKey(1)
        if keypress == ord("q"):
            break
        if keypress == ord("c"):
            if flag_start_capturing == False:
                flag_start_capturing = True
            else:
                flag_start_capturing = False
                frames = 0
        if flag_start_capturing == True:
            frames += 1


if __name__ == "__main__":
    Camera = start_program()
    if Camera == True:
        print("Camera is on")
        Capture_image = check_face(Camera)
    else:
        exit()
    if Capture_image == True:
        init_create_folder()
        img_name = input("Enter image name/text: ")
        store_images(img_name)
